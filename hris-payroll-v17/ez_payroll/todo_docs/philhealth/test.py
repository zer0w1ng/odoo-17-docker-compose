from openpyxl import Workbook,load_workbook
from copy import copy
from __future__ import print_function
filename='test2.xlsx'
wb=Workbook()
wb=load_workbook(filename=filename)
print(wb.sheetnames)
if 1:ws=wb['DATA'];ws['B1'].value='Euro-CB Phils. Inc.';wb.save('test-out.xlsx')
wb.close()