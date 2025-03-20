from openpyxl import load_workbook
import csv
from pprint import pprint
import json,re,random
from datetime import datetime
from dateutil.relativedelta import relativedelta
from string import ascii_lowercase
import urllib.request,odoorpc
def import_xlsx(alphalist_id,xls_file='NTI FINAL ALPHALIST 2024.xlsx',id_prefix='__import__.sch1_2024_%s'):
	F=load_workbook(xls_file,read_only=True,data_only=True);G=F['Sheet1'];B=17;J=B;H=['id','alphalist_id','nationality','employment_status','hired','date_end','pres_nt_income','pres_nt_13mp','pres_nt_deminimis','pres_nt_govded','pres_nt_others','pres_t_income','pres_t_13mp','pres_t_others','tin_no','pres_tax_withheld','tax_paid','tax_refunded','prev_nt_13mp','prev_nt_deminimis','prev_nt_govded','prev_t_others','prev_t_income','prev_t_13mp','prev_t_others','prev_tax_withheld'];C=[];D=B
	for A in G.iter_rows(min_row=B):
		I=(A[1].value or'').strip()
		if I:E=[id_prefix%D,alphalist_id,(A[2].value or'').strip(),(A[3].value or'').strip(),A[4].value.strftime('%Y-%m-%d')or'',A[5].value.strftime('%Y-%m-%d')or'',.0,A[8].value or'',A[9].value or'',A[10].value or'',A[11].value or'',A[13].value or'',A[14].value or'',A[15].value or'',(A[17].value or'').strip(),A[35].value or'',A[37].value or'',A[38].value or'',A[23].value or'',A[24].value or'',A[25].value or'',A[26].value or'',A[28].value or'',A[29].value or'',A[30].value or'',A[34].value or''];print(E);C.append(E)
		D+=1
	return H,C
def odoo_load(odoo,table,header,data):A=table;print();print('Load Data:',A);odoo.env.context.update({'tracking_disable':True});B=odoo.env[A];print('load records=%d'%len(data[1:]));C=B.load(header,data);print('res=%s'%C)
def capitalize_name(odoo):
	B=odoo.env['hr.employee'].search(('|',['active','=',True],['active','=',False]));print(B)
	for A in odoo.env['hr.employee'].browse(B):
		if any(A.islower()for A in A.name):
			if 1:
				if A.last_name:A.last_name=A.last_name.upper()
				if A.first_name:A.first_name=A.first_name.upper()
				if A.middle_name:A.middle_name=A.middle_name.upper()
			print(A.name)
if __name__=='__main__':
	host='novare.eztechsoft.com';port=443;db='novare';user='rberdin@eztechsoft.com';pwd='4660d59dc61fd106146a842349a1e3024cd815ee';protocol='jsonrpc+ssl';odoo=odoorpc.ODOO(host,port=port,protocol=protocol);odoo.login(db,user,pwd)
	if 0:capitalize_name(odoo)
	if 0:
		table='ez.bir.alphalist';alphalist='Novare 2024';header,data=import_xlsx(alphalist)
		if 0:
			for d in data[200:]:
				name=d[2].strip();rec=odoo.env['hr.employee'].search([('name','=',name),'|',('active','=',True),('active','=',False)]);print(name,rec)
				if not rec:break
		table='ez.bir.alphalist.schedule1'
		if 0:
			alphalist_id=odoo.env['ez.bir.alphalist'].search([('name','=',alphalist)])[0];print(alphalist_id);rec_ids=odoo.env['ez.bir.alphalist.schedule1'].search([('alphalist_id','=',alphalist_id)]);print(rec_ids)
			for rec in odoo.env['ez.bir.alphalist.schedule1'].browse(rec_ids):sname=rec.pres_t_others_text;emp_id=odoo.env['hr.employee'].search((['name','=',rec.pres_t_others_text],'|',['active','=',True],['active','=',False]))[0];rec.employee_id=emp_id
		if 0:alphalist_id=odoo.env['ez.bir.alphalist'].search([('name','=',alphalist)])[0];rec_ids=odoo.env['ez.bir.alphalist.schedule1'].search([('alphalist_id','=',alphalist_id),('date_end','!=','2024-12-31')]);odoo.env['ez.bir.alphalist.schedule1'].write(rec_ids,{'separation_reason':'T'})
		if 0:alphalist_id=odoo.env['ez.bir.alphalist'].search([('name','=',alphalist)])[0];rec_ids=odoo.env['ez.bir.alphalist.schedule1'].search([('alphalist_id','=',alphalist_id)]);new_text='Leaves/Night Diff./Adj.';odoo.env['ez.bir.alphalist.schedule1'].write(rec_ids,{'pres_t_others_text':new_text})
		if 0:
			alphalist_id=odoo.env['ez.bir.alphalist'].search([('name','=',alphalist)])[0];recs=odoo.env['ez.bir.alphalist.schedule1'].search_read([('alphalist_id','=',alphalist_id)],['employee_id','tax_due','pres_tax_withheld','prev_tax_withheld','tax_paid','tax_refunded'])
			for rec in recs:
				tax=round(rec['tax_due']-rec['pres_tax_withheld']-rec['prev_tax_withheld'],2)
				if tax>0:
					if rec['tax_paid']!=tax:print('paid',rec['employee_id'][1],tax,rec['tax_paid']);arec=odoo.env['ez.bir.alphalist.schedule1'].browse(rec['id']);arec.tax_paid=tax
				elif rec['tax_refunded']!=-tax:print('refu',rec['employee_id'][1],-tax,rec['tax_refunded']);arec=odoo.env['ez.bir.alphalist.schedule1'].browse(rec['id']);arec.tax_refunded=-tax
	if 1:
		table='ez.bir.alphalist';alphalist='MDI 2024';header,data=import_xlsx(alphalist,xls_file='MDI FINAL ALPHALIST 2024.xlsx',id_prefix='__import__.mdi_sch1_2024_%s')
		if 0:
			for d in data[400:]:
				name=d[2].strip();rec=odoo.env['hr.employee'].search([('name','=',name),'|',('active','=',True),('active','=',False)]);print(name,rec)
				if not rec:break
		table='ez.bir.alphalist.schedule1'
		if 0:
			alphalist_id=odoo.env['ez.bir.alphalist'].search([('name','=',alphalist)])[0];print(alphalist_id);rec_ids=odoo.env['ez.bir.alphalist.schedule1'].search([('alphalist_id','=',alphalist_id)]);print(rec_ids)
			for rec in odoo.env['ez.bir.alphalist.schedule1'].browse(rec_ids):sname=rec.pres_t_others_text;emp_id=odoo.env['hr.employee'].search((['name','=',rec.pres_t_others_text],'|',['active','=',True],['active','=',False]))[0];rec.employee_id=emp_id
		if 1:alphalist_id=odoo.env['ez.bir.alphalist'].search([('name','=',alphalist)])[0];rec_ids=odoo.env['ez.bir.alphalist.schedule1'].search([('alphalist_id','=',alphalist_id),('date_end','!=','2024-12-31')]);odoo.env['ez.bir.alphalist.schedule1'].write(rec_ids,{'separation_reason':'T'})
		if 1:alphalist_id=odoo.env['ez.bir.alphalist'].search([('name','=',alphalist)])[0];rec_ids=odoo.env['ez.bir.alphalist.schedule1'].search([('alphalist_id','=',alphalist_id)]);new_text='Leaves/Night Diff./Adj.';odoo.env['ez.bir.alphalist.schedule1'].write(rec_ids,{'pres_t_others_text':new_text})
		if 0:
			alphalist_id=odoo.env['ez.bir.alphalist'].search([('name','=',alphalist)])[0];recs=odoo.env['ez.bir.alphalist.schedule1'].search_read([('alphalist_id','=',alphalist_id)],['employee_id','tax_due','pres_tax_withheld','prev_tax_withheld','tax_paid','tax_refunded'])
			for rec in recs:
				tax=round(rec['tax_due']-rec['pres_tax_withheld']-rec['prev_tax_withheld'],2)
				if tax>0:
					if rec['tax_paid']!=tax:print('paid',rec['employee_id'][1],tax,rec['tax_paid']);arec=odoo.env['ez.bir.alphalist.schedule1'].browse(rec['id']);arec.tax_paid=tax
				elif rec['tax_refunded']!=-tax:print('refu',rec['employee_id'][1],-tax,rec['tax_refunded']);arec=odoo.env['ez.bir.alphalist.schedule1'].browse(rec['id']);arec.tax_refunded=-tax